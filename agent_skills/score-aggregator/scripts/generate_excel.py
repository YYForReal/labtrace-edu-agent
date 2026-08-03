#!/usr/bin/env python3
"""
Excel 报表生成脚本
生成四工作表的成绩统计 Excel 文件
"""

import sys
import json
import argparse
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def generate_excel(matched_data, statistics, output_path, rubric=None):
    """
    生成 Excel 成绩报表

    Args:
        matched_data: 匹配后的成绩数据
        statistics: 统计数据
        output_path: 输出 Excel 文件路径
        rubric: 评分标准 dict（可选），提供时会展开分项成绩列
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)

    records = (
        matched_data.get("all_records", matched_data)
        if isinstance(matched_data, dict)
        else matched_data
    )

    # 1. 成绩总表
    _create_score_sheet(wb, records, rubric)

    # 2. 统计汇总
    _create_stats_sheet(wb, statistics)

    # 3. 分数分布
    _create_distribution_sheet(wb, statistics)

    # 4. 未提交学生
    unmatched = [r for r in records if r.get("状态") == "未提交"]
    if unmatched:
        _create_unmatched_sheet(wb, unmatched)

    wb.save(output_path)
    return output_path


# 样式常量
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
RED_FILL = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _create_score_sheet(wb, records, rubric=None):
    """成绩总表（含分项成绩列）"""
    ws = wb.create_sheet("成绩总表")

    # 按总分降序排序
    sorted_records = sorted(records, key=lambda r: r.get("总分", 0), reverse=True)

    # 从 rubric 提取子项列定义
    criteria_columns = []
    if rubric:
        criteria_columns = [
            {"id": c["id"], "name": c["name"], "max_score": c["max_score"]}
            for c in rubric.get("criteria", [])
        ]

    # 如果没有 rubric，尝试从第一条记录的 criterion_scores 推断
    if not criteria_columns:
        for r in sorted_records:
            cs_list = r.get("criterion_scores", [])
            if cs_list:
                criteria_columns = [
                    {
                        "id": cs.get("criterion_id", ""),
                        "name": cs.get("criterion_name", ""),
                        "max_score": cs.get("max_score", 0),
                    }
                    for cs in cs_list
                ]
                break

    # 构建表头：学号 | 姓名 | [各子项] | 总分 | 等级 | 状态
    headers = ["学号", "姓名"]
    for c in criteria_columns:
        headers.append(f"{c['name']}({c['max_score']})")
    headers.extend(["总分", "等级", "状态"])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # criterion_id → 列索引 映射
    cid_to_col = {}
    for idx, c in enumerate(criteria_columns):
        cid_to_col[c["id"]] = idx + 3

    total_col = len(criteria_columns) + 3
    grade_col = len(criteria_columns) + 4
    status_col = len(criteria_columns) + 5

    # 数据行
    for row_idx, record in enumerate(sorted_records, 2):
        score = record.get("总分", 0)
        is_graded = record.get("状态") == "已批改"

        # 等级判定
        if score >= 90:
            grade = "A"
        elif score >= 80:
            grade = "B"
        elif score >= 70:
            grade = "C"
        elif score >= 60:
            grade = "D"
        else:
            grade = "F"

        # 学号、姓名
        ws.cell(row=row_idx, column=1, value=record.get("学号", "")).border = (
            THIN_BORDER
        )
        ws.cell(row=row_idx, column=2, value=record.get("姓名", "")).border = (
            THIN_BORDER
        )

        # 各子项分数
        for cs in record.get("criterion_scores", []):
            col = cid_to_col.get(cs.get("criterion_id", ""))
            if col:
                cell = ws.cell(
                    row=row_idx,
                    column=col,
                    value=cs.get("score", 0) if is_graded else "",
                )
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER
                # 扣分原因写入批注
                reason = cs.get("reason", "")
                if reason and is_graded:
                    try:
                        from openpyxl.comments import Comment

                        cell.comment = Comment(reason, "AI批改助手")
                        cell.comment.width = 300
                        cell.comment.height = 100
                    except Exception:
                        pass

        # 未填写的子项列也加边框
        for c in criteria_columns:
            col = cid_to_col[c["id"]]
            cell = ws.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER
            if cell.value is None:
                cell.value = ""
            cell.alignment = CENTER_ALIGN

        # 总分、等级、状态
        total_cell = ws.cell(
            row=row_idx, column=total_col, value=score if is_graded else ""
        )
        total_cell.alignment = CENTER_ALIGN
        total_cell.border = THIN_BORDER

        grade_cell = ws.cell(
            row=row_idx, column=grade_col, value=grade if is_graded else ""
        )
        grade_cell.alignment = CENTER_ALIGN
        grade_cell.border = THIN_BORDER

        status_cell = ws.cell(
            row=row_idx, column=status_col, value=record.get("状态", "")
        )
        status_cell.alignment = CENTER_ALIGN
        status_cell.border = THIN_BORDER

    # 列宽
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    from openpyxl.utils import get_column_letter

    for c in criteria_columns:
        col_letter = get_column_letter(cid_to_col[c["id"]])
        ws.column_dimensions[col_letter].width = min(
            20, max(10, len(c["name"]) * 2 + 6)
        )
    ws.column_dimensions[get_column_letter(total_col)].width = 10
    ws.column_dimensions[get_column_letter(grade_col)].width = 8
    ws.column_dimensions[get_column_letter(status_col)].width = 10

    ws.freeze_panes = "A2"


def _create_stats_sheet(wb, stats):
    """统计汇总"""
    ws = wb.create_sheet("统计汇总")

    data = [
        ["统计项目", "数值"],
        ["班级总人数", stats.get("total_students", 0)],
        ["提交人数", stats.get("submitted_students", 0)],
        ["已批改人数", stats.get("graded_students", 0)],
        ["", ""],
        ["平均分", stats.get("average_score", 0)],
        ["最高分", stats.get("max_score", 0)],
        ["最低分", stats.get("min_score", 0)],
        ["中位数", stats.get("median_score", 0)],
        ["标准差", stats.get("std_deviation", 0)],
        ["", ""],
        ["及格率", f"{stats.get('pass_rate', 0)}%"],
        ["优秀率(≥90)", f"{stats.get('excellent_rate', 0)}%"],
        ["良好率(80-89)", f"{stats.get('good_rate', 0)}%"],
        ["中等率(70-79)", f"{stats.get('fair_rate', 0)}%"],
        ["不及格率(<60)", f"{stats.get('poor_rate', 0)}%"],
    ]

    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

            if row_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = CENTER_ALIGN
            elif col_idx == 2:
                cell.alignment = CENTER_ALIGN

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15


def _create_distribution_sheet(wb, stats):
    """分数分布"""
    ws = wb.create_sheet("分数分布")

    dist = stats.get("score_distribution", {})
    submitted = stats.get("submitted_students", 1) or 1

    data = [
        ["分数段", "人数", "占比"],
        [
            "90-100",
            dist.get("90-100", 0),
            f"{dist.get('90-100', 0)/submitted*100:.1f}%",
        ],
        ["80-89", dist.get("80-89", 0), f"{dist.get('80-89', 0)/submitted*100:.1f}%"],
        ["70-79", dist.get("70-79", 0), f"{dist.get('70-79', 0)/submitted*100:.1f}%"],
        ["60-69", dist.get("60-69", 0), f"{dist.get('60-69', 0)/submitted*100:.1f}%"],
        ["0-59", dist.get("0-59", 0), f"{dist.get('0-59', 0)/submitted*100:.1f}%"],
    ]

    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

            if row_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10

    # 柱状图
    chart = BarChart()
    chart.title = "分数分布直方图"
    chart.x_axis.title = "分数段"
    chart.y_axis.title = "人数"
    chart.style = 10

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=6)
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4

    ws.add_chart(chart, "E2")


def _create_unmatched_sheet(wb, records):
    """未提交学生"""
    ws = wb.create_sheet("未提交学生")

    headers = ["学号", "姓名", "状态"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = RED_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, record in enumerate(records, 2):
        values = [record.get("学号", ""), record.get("姓名", ""), "未提交"]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10


def main():
    parser = argparse.ArgumentParser(description="生成成绩 Excel 报表")
    parser.add_argument("--matched", "-m", required=True, help="匹配后的成绩 JSON")
    parser.add_argument("--stats", "-s", required=True, help="统计数据 JSON")
    parser.add_argument("--output", "-o", required=True, help="输出 Excel 文件路径")
    parser.add_argument(
        "--rubric", "-r", help="评分标准 JSON 文件路径（可选，提供时展开分项成绩列）"
    )

    args = parser.parse_args()

    with open(args.matched, "r", encoding="utf-8") as f:
        matched_data = json.load(f)

    with open(args.stats, "r", encoding="utf-8") as f:
        statistics = json.load(f)

    rubric = None
    if args.rubric:
        with open(args.rubric, "r", encoding="utf-8") as f:
            rubric = json.load(f)

    try:
        output = generate_excel(matched_data, statistics, args.output, rubric=rubric)
        print(f"Excel 报表已生成: {output}")
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
