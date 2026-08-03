#!/usr/bin/env python3
"""
fill_score_to_excel.py — 成绩回填到学校成绩登记表

功能概述：
  将 AI 批改得到的成绩数据回填到学校统一格式的《学生成绩登记表》Excel 中。
  支持按学号匹配、自动定位成绩列、保留原始表格格式和合并单元格。

学校成绩登记表结构（深圳大学标准）：
  Row 1:   （空或标题装饰行）
  Row 2:   "深圳大学学生成绩登记表"（合并单元格 A2:V2）
  Row 3:   "YYYY-YYYY学年第X学期"（合并单元格 A3:V3）
  Row 4:   日期 + 总人数
  Row 5-6: 课程信息（编号、班号、开课单位、课程名、类别、学分、主讲教师）
  Row 7:   表头第1行：序号 | 学号 | 姓名 | 性别 | 主修专业 | 平时考核方式与结果 | 平时总评成绩 | 期末基本题成绩 | 期末附加题成绩 | 备注
  Row 8:   表头第2行：                              1 | 2 | ... | 11（考核周次）
  Row 9:   表头第3行（可能为空）
  Row 10+: 学生数据行

关键列定位策略：
  - 学号列：搜索表头行中包含"学号"的列
  - 姓名列：搜索表头行中包含"姓名"的列
  - 平时考核列 H-R：对应每周实验/考勤成绩
  - 平时总评成绩列 S：回填实验报告总分
  - 期末基本题成绩列 T、期末附加题成绩列 U

用法：
  python fill_score_to_excel.py \
    --excel 成绩登记表.xlsx \
    --scores grading_results.json \
    --output 成绩登记表_已填.xlsx \
    [--score-column S] \
    [--week-column H] \
    [--week-number 5]

  或在 Python 中导入：
    from fill_score_to_excel import fill_scores_to_excel
"""

import argparse
import copy
import json
import os
import re
import shutil
import sys
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def _safe_save(wb, output_path):
    """安全保存工作簿（通过临时文件避免 macOS 文件锁冲突）"""
    import tempfile

    out_dir = os.path.dirname(os.path.abspath(output_path)) or "."
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=out_dir)
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        shutil.move(tmp_path, output_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        wb.save(output_path)


def _build_comment_text(score_item):
    """
    构建 Excel 单元格批注文本。

    将评语（comment）和分项成绩（criterion_scores / scores）组合为格式清晰的多行文本，
    悬停单元格即可查看完整的批改说明。

    Args:
        score_item: dict — 成绩数据项

    Returns:
        str|None — 批注文本，无内容时返回 None
    """
    parts = []

    total = score_item.get("total_score", 0)

    # 优先使用 criterion_scores（含子项名称和扣分原因）
    criterion_scores = score_item.get("criterion_scores", [])
    if criterion_scores:
        lines = []
        for cs in criterion_scores:
            name = cs.get("criterion_name", cs.get("criterion_id", ""))
            score = cs.get("score", 0)
            max_s = cs.get("max_score", 0)
            reason = cs.get("reason", "")
            line = f"  {name}: {score}/{max_s}"
            if reason:
                line += f" — {reason}"
            lines.append(line)
        parts.append(f"【分项明细】总分 {total}")
        parts.extend(lines)
    else:
        # 回退：使用 scores 数字数组
        scores = score_item.get("scores", [])
        if scores:
            score_formula = "+".join(str(s) for s in scores) + f"={total}"
            parts.append(f"【成绩明细】{score_formula}")

    # 等级（如有）
    grade = score_item.get("grade", "")
    if grade:
        parts.append(f"【等级】{grade}")

    # 评语
    comment = score_item.get("comment", "")
    if comment:
        parts.append(f"【评语】{comment}")

    # 批注内容（annotations 摘要，如有）
    annotations = score_item.get("annotations", [])
    if annotations:
        parts.append(f"【批注数】{len(annotations)} 条")

    if not parts:
        return None

    return "\n".join(parts)


# ══════════════════════════════════════════════════════════
# 表格结构自动检测
# ══════════════════════════════════════════════════════════


def _detect_table_structure(ws):
    """
    自动检测成绩登记表的结构。

    返回 dict:
      header_row: int — 表头行号（包含"学号""姓名"的行）
      data_start_row: int — 数据起始行号
      data_end_row: int — 数据结束行号
      col_map: dict — 列名到列号的映射
        {
          '序号': 1, '学号': 2, '姓名': 3, '性别': 4, '主修专业': 5,
          '平时考核_start': 8, '平时考核_end': 18,
          '平时总评成绩': 19, '期末基本题成绩': 20, '期末附加题成绩': 21, '备注': 22
        }
      students: list[dict] — 学生列表 [{row, sid, name}, ...]
    """
    structure = {
        "header_row": None,
        "data_start_row": None,
        "data_end_row": None,
        "col_map": {},
        "students": [],
    }

    # 1. 定位表头行：搜索包含"学号"和"姓名"的行
    for row_idx in range(1, min(20, ws.max_row + 1)):
        row_texts = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                row_texts[col_idx] = str(val).strip()

        has_sid = any("学号" in v for v in row_texts.values())
        has_name = any("姓名" in v for v in row_texts.values())

        if has_sid and has_name:
            structure["header_row"] = row_idx

            # 映射列
            for col_idx, text in row_texts.items():
                if "序号" in text:
                    structure["col_map"]["序号"] = col_idx
                elif "学号" == text or text == "学号":
                    structure["col_map"]["学号"] = col_idx
                elif "姓名" == text or text == "姓名":
                    structure["col_map"]["姓名"] = col_idx
                elif "性别" in text:
                    structure["col_map"]["性别"] = col_idx
                elif "主修专业" in text or "专业" in text:
                    structure["col_map"]["主修专业"] = col_idx
                elif "平时考核" in text:
                    structure["col_map"]["平时考核_header"] = col_idx
                elif "平时总评" in text:
                    structure["col_map"]["平时总评成绩"] = col_idx
                elif "期末基本题" in text:
                    structure["col_map"]["期末基本题成绩"] = col_idx
                elif "期末附加题" in text:
                    structure["col_map"]["期末附加题成绩"] = col_idx
                elif "备注" in text:
                    structure["col_map"]["备注"] = col_idx
            break

    if structure["header_row"] is None:
        print("警告：无法自动检测表头行", file=sys.stderr)
        return structure

    # 2. 检测平时考核子列（周次编号行，通常在 header_row + 1）
    sub_header_row = structure["header_row"] + 1
    week_cols = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=sub_header_row, column=col_idx).value
        if val is not None:
            try:
                week_num = int(float(val))
                if 1 <= week_num <= 20:
                    week_cols.append((col_idx, week_num))
            except (ValueError, TypeError):
                pass

    if week_cols:
        structure["col_map"]["平时考核_start"] = week_cols[0][0]
        structure["col_map"]["平时考核_end"] = week_cols[-1][0]
        structure["col_map"]["week_columns"] = {wn: col for col, wn in week_cols}

    # 3. 定位数据区域
    sid_col = structure["col_map"].get("学号", 2)
    name_col = structure["col_map"].get("姓名", 3)

    # 数据从表头后第2-3行开始（跳过子表头）
    search_start = structure["header_row"] + 2
    for row_idx in range(search_start, ws.max_row + 1):
        sid_val = ws.cell(row=row_idx, column=sid_col).value
        name_val = ws.cell(row=row_idx, column=name_col).value
        if sid_val is not None and name_val is not None:
            sid_str = str(sid_val).strip()
            # 验证学号格式（纯数字，>=6位）
            if re.match(r"^\d{6,}$", sid_str.replace(".0", "")):
                if structure["data_start_row"] is None:
                    structure["data_start_row"] = row_idx
                structure["data_end_row"] = row_idx
                structure["students"].append(
                    {
                        "row": row_idx,
                        "sid": sid_str.replace(".0", ""),
                        "name": str(name_val).strip(),
                    }
                )

    return structure


# ══════════════════════════════════════════════════════════
# 成绩回填
# ══════════════════════════════════════════════════════════


def fill_scores_to_excel(
    excel_path,
    scores_data,
    output_path,
    score_column=None,
    week_column=None,
    week_number=None,
    fill_mode="total_score",
    add_comments=True,
    comment_author="AI批改助手",
    auto_stats=True,
):
    """
    将批改成绩回填到学校成绩登记表。

    Args:
        excel_path: str — 成绩登记表 Excel 路径
        scores_data: list[dict] — 成绩数据列表，每个 dict 至少包含：
            {
                "student_id": "2022150022",
                "student_name": "林振法",      # 可选，用于辅助匹配和校验
                "total_score": 83,             # 总分
                "scores": [37, 3, 3, 5, 7, 5, 23],  # 可选，各维度分项
                "comment": "评语文本",          # 可选，写入 Excel 单元格批注
                "grade": "B"                    # 可选，等级
            }
        output_path: str — 输出 Excel 路径
        score_column: str|int|None — 手动指定回填的目标列（如 'S' 或 19）
            为 None 时自动检测"平时总评成绩"列
        week_column: str|int|None — 手动指定周次考核列（如 'L' 或 12）
        week_number: int|None — 写入哪一周的考核列
        fill_mode: str — 填充模式：
            'total_score' — 仅填写总分到指定列（默认）
            'week_score' — 填写到某一周的考核列
            'both' — 同时填写周次列和总评列
        add_comments: bool — 是否在成绩单元格上添加批注说明（默认 True）
            批注内容来自 scores_data 中的 "comment" 字段，包含评语和分项成绩明细
        comment_author: str — 批注的作者名（默认 'AI批改助手'）
        auto_stats: bool — 是否自动追加统计汇总和分数分布工作表（默认 True）

    Returns:
        dict: 回填结果
            {
                'matched': int,         # 成功匹配并填写的学生数
                'unmatched_scores': [],  # 有成绩但在表中找不到的学号
                'unfilled_students': [], # 在表中但没有成绩的学生
                'name_mismatches': [],   # 学号匹配但姓名不一致的警告
                'comments_added': int,   # 成功添加的批注数
                'total_in_excel': int,   # 表中总学生数
                'total_scores': int,     # 传入的成绩数
            }
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("请安装 openpyxl: pip install openpyxl")

    # 加载工作簿（保留样式和合并单元格）
    wb = load_workbook(excel_path)
    ws = wb.active

    # 自动检测表格结构
    structure = _detect_table_structure(ws)
    print(
        f"检测到 {len(structure['students'])} 个学生，"
        f"数据行 {structure['data_start_row']}-{structure['data_end_row']}"
    )

    # 确定目标列
    if score_column is not None:
        if isinstance(score_column, str):
            target_col = column_index_from_string(score_column)
        else:
            target_col = int(score_column)
    else:
        target_col = structure["col_map"].get("平时总评成绩")
        if target_col is None:
            # 回退：平时考核最后一列 + 1
            pk_end = structure["col_map"].get("平时考核_end")
            if pk_end:
                target_col = pk_end + 1
            else:
                target_col = 19  # 默认 S 列

    print(f"目标列: {get_column_letter(target_col)} (列号 {target_col})")

    # 确定周次列（如有需要）
    week_target_col = None
    if week_number is not None and fill_mode in ("week_score", "both"):
        if week_column is not None:
            if isinstance(week_column, str):
                week_target_col = column_index_from_string(week_column)
            else:
                week_target_col = int(week_column)
        else:
            week_cols = structure["col_map"].get("week_columns", {})
            week_target_col = week_cols.get(week_number)

        if week_target_col:
            print(f"周次列: {get_column_letter(week_target_col)} (第 {week_number} 周)")

    # 构建学号到成绩的映射
    score_map = {}
    for item in scores_data:
        sid = str(item.get("student_id", "")).strip()
        if sid:
            # 标准化学号（去掉 .0 后缀）
            sid = sid.replace(".0", "")
            score_map[sid] = item

    # 执行回填
    result = {
        "matched": 0,
        "unmatched_scores": [],
        "unfilled_students": [],
        "name_mismatches": [],
        "comments_added": 0,
        "total_in_excel": len(structure["students"]),
        "total_scores": len(scores_data),
        "details": [],
    }

    matched_sids = set()

    for student in structure["students"]:
        sid = student["sid"]
        row = student["row"]
        name = student["name"]

        if sid in score_map:
            matched_sids.add(sid)
            score_item = score_map[sid]
            total_score = score_item.get("total_score", 0)

            # 姓名校验
            expected_name = score_item.get("student_name", "")
            if expected_name and expected_name != name:
                result["name_mismatches"].append(
                    {
                        "sid": sid,
                        "excel_name": name,
                        "score_name": expected_name,
                        "row": row,
                    }
                )
                print(f"  ⚠ 姓名不一致: {sid} 表中={name} vs 成绩={expected_name}")

            # 写入总分
            if fill_mode in ("total_score", "both"):
                ws.cell(row=row, column=target_col, value=total_score)

            # 写入周次分数
            if week_target_col and fill_mode in ("week_score", "both"):
                ws.cell(row=row, column=week_target_col, value=total_score)

            # 添加批注说明
            if add_comments:
                comment_text = _build_comment_text(score_item)
                if comment_text:
                    # 批注附加在总评成绩列的单元格上
                    comment_cell_col = (
                        target_col
                        if fill_mode in ("total_score", "both")
                        else week_target_col
                    )
                    if comment_cell_col:
                        cell = ws.cell(row=row, column=comment_cell_col)
                        cell.comment = Comment(comment_text, comment_author)
                        # 设置批注框尺寸（宽度和高度，单位为像素）
                        cell.comment.width = 350
                        cell.comment.height = 200
                        result["comments_added"] += 1

            result["matched"] += 1
            result["details"].append(
                {
                    "sid": sid,
                    "name": name,
                    "score": total_score,
                    "row": row,
                    "status": "filled",
                }
            )
            print(f"  ✓ {sid} {name}: {total_score}")
        else:
            result["unfilled_students"].append(
                {
                    "sid": sid,
                    "name": name,
                    "row": row,
                }
            )

    # 检查有成绩但不在表中的学生
    for sid, item in score_map.items():
        if sid not in matched_sids:
            result["unmatched_scores"].append(
                {
                    "sid": sid,
                    "name": item.get("student_name", ""),
                    "score": item.get("total_score", 0),
                }
            )
            print(f"  ✗ {sid} {item.get('student_name', '')}: 不在登记表中")

    # 保存
    _safe_save(wb, output_path)
    print(f"\n成绩回填完成: {output_path}")
    print(
        f"  匹配 {result['matched']}/{result['total_in_excel']} 人, "
        f"未匹配成绩 {len(result['unmatched_scores'])} 人, "
        f"未填写 {len(result['unfilled_students'])} 人"
    )
    if result["comments_added"] > 0:
        print(f"  批注说明 {result['comments_added']} 条")

    # 自动追加统计工作表
    if auto_stats and scores_data:
        print(f"正在追加统计工作表...")
        append_statistics_sheet(output_path, scores_data)

    return result


# ══════════════════════════════════════════════════════════
# 从批改配置中提取成绩数据
# ══════════════════════════════════════════════════════════


def extract_scores_from_grading_configs(config_dir, file_pattern=None):
    """
    从批改配置目录中批量提取成绩数据。

    Args:
        config_dir: str — 批改配置 JSON 文件目录
        file_pattern: str|None — 文件名匹配模式（正则），
            默认匹配 "学号+姓名.json" 格式

    Returns:
        list[dict] — 成绩数据列表
    """
    scores = []

    if file_pattern is None:
        # 默认匹配：文件名中包含学号
        file_pattern = r"(\d{10})\s*[_\-]?\s*(.+)\.json$"

    for fname in sorted(os.listdir(config_dir)):
        if not fname.endswith(".json"):
            continue

        fpath = os.path.join(config_dir, fname)
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                config = json.load(f)
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            print(f"  跳过无效文件: {fname} ({e})", file=sys.stderr)
            continue

        # 提取学号和姓名
        sid = config.get("student_id", "")
        name = config.get("student_name", "")

        # 从文件名提取学号和姓名（如果配置中没有）
        if not sid:
            m = re.search(r"(\d{10})", fname)
            if m:
                sid = m.group(1)
        if not name:
            m = re.search(r"\d{10}\s*[_\-]?\s*(.+?)\.json$", fname)
            if m:
                name = m.group(1).strip()

        # 提取分数
        score_list = config.get("scores", [])
        total = sum(score_list) if score_list else config.get("total_score", 0)

        if sid and total > 0:
            scores.append(
                {
                    "student_id": sid,
                    "student_name": name,
                    "total_score": total,
                    "scores": score_list,
                    "comment": config.get("comment", ""),
                }
            )

    return scores


def extract_scores_from_results_json(results_path):
    """
    从批量批改结果 JSON 中提取成绩数据。

    JSON 格式：
    [
        {
            "student_id": "2022150022",
            "student_name": "林振法",
            "total_score": 83,
            "scores": [37, 3, 3, 5, 7, 5, 23],
            "comment": "..."
        },
        ...
    ]
    """
    with open(results_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        return data
    elif isinstance(data, dict):
        # 兼容 {results: [...]} 格式
        return data.get("results", data.get("grading_results", [data]))
    return []


# ══════════════════════════════════════════════════════════
# 统计信息附加（在已有工作簿上追加统计工作表）
# ══════════════════════════════════════════════════════════


def append_statistics_sheet(excel_path, scores_data, output_path=None):
    """
    在成绩登记表工作簿中追加统计汇总和分数分布工作表。
    不影响原有的成绩数据表。

    Args:
        excel_path: str — 已回填成绩的 Excel 路径
        scores_data: list[dict] — 成绩数据列表
        output_path: str|None — 输出路径，None 则覆盖原文件
    """
    if output_path is None:
        output_path = excel_path

    wb = load_workbook(excel_path)

    # 提取已有分数
    all_scores = [
        s.get("total_score", 0) for s in scores_data if s.get("total_score", 0) > 0
    ]
    n = len(all_scores)

    if n == 0:
        print("无有效成绩数据，跳过统计", file=sys.stderr)
        _safe_save(wb, output_path)
        return

    # 计算统计
    avg = sum(all_scores) / n
    max_s = max(all_scores)
    min_s = min(all_scores)

    sorted_s = sorted(all_scores)
    median = (
        (sorted_s[n // 2 - 1] + sorted_s[n // 2]) / 2
        if n % 2 == 0
        else sorted_s[n // 2]
    )

    import math

    std = math.sqrt(sum((s - avg) ** 2 for s in all_scores) / (n - 1)) if n > 1 else 0

    dist = {
        "90-100 (A)": sum(1 for s in all_scores if 90 <= s <= 100),
        "80-89  (B)": sum(1 for s in all_scores if 80 <= s < 90),
        "70-79  (C)": sum(1 for s in all_scores if 70 <= s < 80),
        "60-69  (D)": sum(1 for s in all_scores if 60 <= s < 70),
        "0-59   (F)": sum(1 for s in all_scores if s < 60),
    }
    pass_count = sum(1 for s in all_scores if s >= 60)

    # 删除旧的统计表（如有）
    for sheet_name in ["统计汇总", "分数分布"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    # ── 统计汇总表 ──
    ws_stats = wb.create_sheet("统计汇总")

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    CENTER = Alignment(horizontal="center", vertical="center")
    THIN = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    stats_rows = [
        ["统计项目", "数值"],
        ["批改总人数", n],
        ["平均分", round(avg, 2)],
        ["最高分", max_s],
        ["最低分", min_s],
        ["中位数", round(median, 2)],
        ["标准差", round(std, 2)],
        ["及格率", f"{round(pass_count / n * 100, 1)}%"],
        ["", ""],
    ]
    for grade_label, count in dist.items():
        pct = round(count / n * 100, 1) if n > 0 else 0
        stats_rows.append([grade_label, f"{count} 人 ({pct}%)"])

    for r_idx, row_data in enumerate(stats_rows, 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_stats.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            cell.alignment = CENTER

    ws_stats.column_dimensions["A"].width = 20
    ws_stats.column_dimensions["B"].width = 18

    # ── 分数分布表 + 柱状图 ──
    ws_dist = wb.create_sheet("分数分布")

    dist_rows = [["分数段", "人数", "占比"]]
    for grade_label, count in dist.items():
        pct = round(count / n * 100, 1) if n > 0 else 0
        dist_rows.append([grade_label.split("(")[0].strip(), count, f"{pct}%"])

    for r_idx, row_data in enumerate(dist_rows, 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_dist.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            cell.alignment = CENTER

    ws_dist.column_dimensions["A"].width = 12
    ws_dist.column_dimensions["B"].width = 10
    ws_dist.column_dimensions["C"].width = 10

    # 柱状图
    try:
        from openpyxl.chart import BarChart, Reference

        chart = BarChart()
        chart.title = "分数分布直方图"
        chart.x_axis.title = "分数段"
        chart.y_axis.title = "人数"
        chart.style = 10

        data_ref = Reference(ws_dist, min_col=2, min_row=1, max_row=len(dist_rows))
        cats = Reference(ws_dist, min_col=1, min_row=2, max_row=len(dist_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4

        ws_dist.add_chart(chart, "E2")
    except Exception as e:
        print(f"  图表创建失败: {e}", file=sys.stderr)

    _safe_save(wb, output_path)
    print(f"统计工作表已追加到: {output_path}")


# ══════════════════════════════════════════════════════════
# CLI 入口
# ══════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="将批改成绩回填到学校成绩登记表 Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 从结果 JSON 回填总评成绩
  python fill_score_to_excel.py \\
    --excel 成绩登记表.xlsx \\
    --scores results.json \\
    --output 成绩登记表_已填.xlsx

  # 从批改配置目录回填，并指定周次列
  python fill_score_to_excel.py \\
    --excel 成绩登记表.xlsx \\
    --config-dir grading_configs/ \\
    --output 成绩登记表_已填.xlsx \\
    --week-number 5 \\
    --fill-mode both

  # 追加统计工作表
  python fill_score_to_excel.py \\
    --excel 成绩登记表_已填.xlsx \\
    --scores results.json \\
    --append-stats
        """,
    )

    parser.add_argument("--excel", "-e", required=True, help="成绩登记表 Excel 路径")
    parser.add_argument(
        "--scores", "-s", help="成绩数据 JSON 文件路径 (list[dict] 格式)"
    )
    parser.add_argument("--config-dir", "-d", help="批改配置 JSON 目录（文件名含学号）")
    parser.add_argument("--output", "-o", help="输出 Excel 路径（默认覆盖原文件）")
    parser.add_argument("--score-column", "-c", help="手动指定回填目标列（如 S 或 19）")
    parser.add_argument(
        "--week-number", "-w", type=int, help="写入哪一周的考核列（周次编号）"
    )
    parser.add_argument("--week-column", help="手动指定周次列（如 L 或 12）")
    parser.add_argument(
        "--fill-mode",
        default="total_score",
        choices=["total_score", "week_score", "both"],
        help="填充模式：total_score(仅总评) / week_score(仅周次) / both(两者)",
    )
    parser.add_argument(
        "--no-comments",
        action="store_true",
        help="不在成绩单元格上添加批注说明（默认添加）",
    )
    parser.add_argument(
        "--comment-author",
        default="AI批改助手",
        help='批注的作者名（默认 "AI批改助手"）',
    )
    parser.add_argument(
        "--no-stats", action="store_true", help="不追加统计汇总工作表（默认自动追加）"
    )
    parser.add_argument(
        "--append-stats",
        action="store_true",
        help="[兼容] 显式追加统计工作表（已改为默认行为）",
    )

    args = parser.parse_args()

    output_path = args.output or args.excel

    # 加载成绩数据
    scores_data = []
    if args.scores:
        scores_data = extract_scores_from_results_json(args.scores)
        print(f"从 JSON 加载 {len(scores_data)} 条成绩数据")
    elif args.config_dir:
        scores_data = extract_scores_from_grading_configs(args.config_dir)
        print(f"从配置目录加载 {len(scores_data)} 条成绩数据")
    else:
        print("错误：请指定 --scores 或 --config-dir", file=sys.stderr)
        sys.exit(1)

    if not scores_data:
        print("错误：未找到有效的成绩数据", file=sys.stderr)
        sys.exit(1)

    # 回填成绩
    result = fill_scores_to_excel(
        excel_path=args.excel,
        scores_data=scores_data,
        output_path=output_path,
        score_column=args.score_column,
        week_column=args.week_column,
        week_number=args.week_number,
        fill_mode=args.fill_mode,
        add_comments=not args.no_comments,
        comment_author=args.comment_author,
        auto_stats=not args.no_stats,
    )

    # [兼容] --append-stats 已改为默认行为；当 --no-stats 且 --append-stats 同时指定时，以 append-stats 为准
    if args.append_stats and args.no_stats:
        append_statistics_sheet(output_path, scores_data)

    # 输出结果摘要
    print(
        f"\n回填结果: {json.dumps(result, ensure_ascii=False, default=str, indent=2)}"
    )
